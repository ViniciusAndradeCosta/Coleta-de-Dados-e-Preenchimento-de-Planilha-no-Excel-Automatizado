"""
Automação Sofascore -> Excel (Libertadores/Brasileirão)
--------------------------------------------------------
v6.1 — Usa API Oculta do Sofascore para garantir precisão máxima de dados, 
       com o caminho do Excel corrigido.

FLUXO:
    1. Você abre o jogo no Sofascore no Opera GX.
    2. Clica no botão "📋 Copiar este jogo".
    3. O script pega a URL, extrai o ID do jogo e consulta a API para dados exatos.

Requisitos:
    pip install curl_cffi openpyxl pyperclip pygetwindow keyboard
"""

import json
import re
import threading
import time
import tkinter as tk
from tkinter import scrolledtext
from datetime import datetime
from pathlib import Path

from curl_cffi import requests as cffi_requests
from openpyxl import load_workbook
import pyperclip

# ============== CONFIGURAÇÕES — EDITE AQUI ==============
CAMINHO_EXCEL = r"C:\Users\costa\Downloads\libertadores_brasileiros_serie_a_2021_2025.xlsx"
NOME_ABA = "Libertadores 2021-2025"
COLUNA_HORARIO = "A"
COLUNA_ESTADIO = "F"
LINHA_INICIAL_BUSCA = 2
# ========================================================


# --------- 1. Pegar URL via clipboard ---------
def pegar_url_clipboard_automatico(log_fn):
    import pygetwindow as gw
    import keyboard

    try:
        clipboard_anterior = pyperclip.paste()
    except Exception:
        clipboard_anterior = ""

    pyperclip.copy("")
    time.sleep(0.1)

    janelas = [w for w in gw.getAllWindows() if "Opera" in (w.title or "")]
    if not janelas:
        raise RuntimeError("Janela do Opera não encontrada.")

    janela = max(janelas, key=lambda w: (w.width or 0) * (w.height or 0))
    log_fn(f"Focando janela: '{janela.title[:50]}'")

    try:
        if janela.isMinimized:
            janela.restore()
        janela.activate()
    except Exception as e:
        log_fn(f"  aviso: {e}")

    time.sleep(0.3)
    keyboard.send("ctrl+l")
    time.sleep(0.2)
    keyboard.send("ctrl+c")
    time.sleep(0.3)
    keyboard.send("esc")

    url = pyperclip.paste().strip()

    try:
        if clipboard_anterior and clipboard_anterior != url:
            pass 
    except Exception:
        pass

    if not url:
        raise RuntimeError(
            "Clipboard vazio. Tente o 'Modo manual'."
        )

    if not url.startswith("http"):
        if "sofascore" in url.lower() or "." in url:
            url = "https://" + url.lstrip("/")
        else:
            raise RuntimeError(f"O que veio do clipboard não é URL: {url[:80]}")

    return url


def pegar_url_clipboard_manual():
    url = pyperclip.paste().strip()
    if not url:
        raise RuntimeError("Clipboard vazio. Copie a URL no Opera antes.")
    if not url.startswith("http"):
        if "." in url:
            url = "https://" + url.lstrip("/")
        else:
            raise RuntimeError(f"Clipboard não tem URL: {url[:80]}")
    return url


# --------- 2. Extrair horário e estádio VIA API ---------
def extrair_dados_sofascore(url):
    if "sofascore.com" not in url:
        raise RuntimeError(f"URL não é do Sofascore: {url}")

    # 1. Extrair o ID exato da partida pela URL
    m = re.search(r'id:(\d+)', url)
    if not m:
        # Tenta pegar caso seja um link no formato antigo
        m = re.search(r'/(\d+)(?:$|#|\?)', url)
        
    if not m:
        raise RuntimeError("Não foi possível encontrar o número de ID do jogo na URL.")
    
    match_id = m.group(1)
    
    # 2. Bater na API direta (Sem erro de pegar dado do time ao invés do jogo)
    api_url = f"https://api.sofascore.com/api/v1/event/{match_id}"

    headers = {
        "Accept-Language": "pt-BR,pt;q=0.9,en-US;q=0.8,en;q=0.7",
        "Origin": "https://www.sofascore.com",
        "Referer": "https://www.sofascore.com/"
    }
    
    resp = cffi_requests.get(api_url, headers=headers, impersonate="chrome", timeout=15)
    
    if resp.status_code != 200:
        raise RuntimeError(f"Erro {resp.status_code} ao consultar a API para o jogo {match_id}.")

    data = resp.json()
    event = data.get("event", {})
    
    # 3. Pegar horário (convertendo timestamp para hora local)
    horario = None
    timestamp = event.get("startTimestamp")
    if timestamp:
        horario = datetime.fromtimestamp(timestamp).strftime("%H:%M")
        
    # 4. Pegar o estádio preciso da partida
    estadio = None
    venue = event.get("venue", {})
    
    stadium_info = venue.get("stadium", {})
    if stadium_info and stadium_info.get("name"):
        estadio = stadium_info.get("name")
    elif venue and venue.get("name"):
        estadio = venue.get("name")

    if not horario and not estadio:
        raise RuntimeError("A API não retornou informações de horário ou estádio para este jogo.")

    return horario, estadio


# --------- 3. Escrever na planilha ---------
def escrever_planilha(horario, estadio):
    caminho = Path(CAMINHO_EXCEL)
    if not caminho.exists():
        raise RuntimeError(f"Planilha não encontrada: {caminho}\nPor favor, verifique a variável CAMINHO_EXCEL.")

    wb = load_workbook(caminho)
    if NOME_ABA not in wb.sheetnames:
        raise RuntimeError(f"Aba '{NOME_ABA}' não existe. Abas: {wb.sheetnames}")

    ws = wb[NOME_ABA]
    linha = LINHA_INICIAL_BUSCA
    while ws[f"{COLUNA_HORARIO}{linha}"].value not in (None, ""):
        linha += 1
        if linha > 10000:
            raise RuntimeError("Mais de 10000 linhas — abortando.")

    if horario:
        ws[f"{COLUNA_HORARIO}{linha}"] = horario
    if estadio:
        ws[f"{COLUNA_ESTADIO}{linha}"] = estadio

    try:
        wb.save(caminho)
    except PermissionError:
        raise RuntimeError(
            "Não consegui salvar o Excel. Feche o arquivo antes de clicar."
        )

    return linha


# --------- 4. Interface flutuante ---------
class JanelaFlutuante:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Sofascore → Excel")
        self.root.geometry("440x420+50+50")
        self.root.attributes("-topmost", True)
        self.root.configure(bg="#1e1e2e")

        tk.Label(
            self.root,
            text="🏆 Sofascore → Planilha",
            font=("Segoe UI", 13, "bold"),
            bg="#1e1e2e",
            fg="#cdd6f4",
        ).pack(pady=(10, 4))

        self.botao_auto = tk.Button(
            self.root,
            text="📋  Copiar este jogo (automático)",
            font=("Segoe UI", 11, "bold"),
            bg="#89b4fa",
            fg="#1e1e2e",
            activebackground="#74c7ec",
            relief="flat",
            cursor="hand2",
            padx=10,
            pady=8,
            command=self.ao_clicar_auto,
        )
        self.botao_auto.pack(pady=4, padx=12, fill="x")

        self.botao_manual = tk.Button(
            self.root,
            text="📎  Usar URL do clipboard (manual)",
            font=("Segoe UI", 10),
            bg="#a6e3a1",
            fg="#1e1e2e",
            activebackground="#94e2d5",
            relief="flat",
            cursor="hand2",
            padx=10,
            pady=6,
            command=self.ao_clicar_manual,
        )
        self.botao_manual.pack(pady=(0, 6), padx=12, fill="x")

        tk.Label(
            self.root,
            text="Modo manual: no Opera dê Ctrl+L → Ctrl+C, depois clique acima.",
            font=("Segoe UI", 8),
            bg="#1e1e2e",
            fg="#7f849c",
            wraplength=400,
        ).pack(pady=(0, 6))

        self.log = scrolledtext.ScrolledText(
            self.root,
            height=13,
            font=("Consolas", 9),
            bg="#11111b",
            fg="#cdd6f4",
            insertbackground="#cdd6f4",
            relief="flat",
        )
        self.log.pack(pady=6, padx=12, fill="both", expand=True)
        self.log.insert(
            "end",
            "Pronto.\n"
            "  • Automático: deixa a aba do jogo ativa no Opera e clica no botão azul.\n"
            "  • Manual: copia a URL no Opera (Ctrl+L, Ctrl+C) e clica no botão verde.\n",
        )
        self.log.configure(state="disabled")

    def logar(self, msg):
        self.log.configure(state="normal")
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log.insert("end", f"[{timestamp}] {msg}\n")
        self.log.see("end")
        self.log.configure(state="disabled")

    def _desabilitar(self):
        self.botao_auto.configure(state="disabled")
        self.botao_manual.configure(state="disabled")

    def _habilitar(self):
        self.botao_auto.configure(state="normal", text="📋  Copiar este jogo (automático)")
        self.botao_manual.configure(state="normal", text="📎  Usar URL do clipboard (manual)")

    def ao_clicar_auto(self):
        self._desabilitar()
        self.botao_auto.configure(text="⏳ Processando...")
        threading.Thread(target=self._processar, args=(True,), daemon=True).start()

    def ao_clicar_manual(self):
        self._desabilitar()
        self.botao_manual.configure(text="⏳ Processando...")
        threading.Thread(target=self._processar, args=(False,), daemon=True).start()

    def _processar(self, automatico):
        try:
            if automatico:
                self.logar("Pegando URL automaticamente (Ctrl+L, Ctrl+C no Opera)...")
                url = pegar_url_clipboard_automatico(self.logar)
            else:
                self.logar("Lendo URL do clipboard...")
                url = pegar_url_clipboard_manual()

            self.logar(f"URL: {url[:80]}{'...' if len(url) > 80 else ''}")
            self.logar("Baixando dados do jogo via API...")
            horario, estadio = extrair_dados_sofascore(url)
            self.logar(f"  Horário: {horario or '(não encontrado)'}")
            self.logar(f"  Estádio: {estadio or '(não encontrado)'}")

            self.logar("Escrevendo na planilha...")
            linha = escrever_planilha(horario, estadio)
            self.logar(f"✅ Gravado na linha {linha}.")
        except Exception as e:
            self.logar(f"❌ ERRO: {e}")
        finally:
            self._habilitar()

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    JanelaFlutuante().run()